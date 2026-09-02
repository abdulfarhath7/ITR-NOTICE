"""The summary report: the firm's old Excel tracker, rebuilt.

Two callers, one set of numbers:
  build_summary(con)   -> what GET /api/summary returns and the dashboard
                          draws: run info, urgency buckets, the attention
                          list and the full register
  build_workbook(data) -> the same dict as a real three-sheet .xlsx

The bucket rules live here and nowhere else. A notice is "open" while its
proceeding is not Closed; only open notices land in an urgency bucket, and a
closed one is counted on its own line, so a notice that has already been
answered can never show up as overdue. Since the scraper started reading the
portal's Submit/View Response button, "answered" means that too: a notice with
a reply already filed is counted as Responded and never as work outstanding.
"""
import io
from datetime import date, datetime

from . import db

# What the portal prints is 17-Aug-2026, and Claude is asked for the same
# shape. The rest are here so one hand-edited row cannot break the report.
DATE_FORMATS = ("%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")

# key -> the words the dashboard and the spreadsheet both use, in this order.
# Every notice lands in exactly one of these.
BUCKETS = [
    ("overdue", "Overdue"),
    ("due_3", "Due ≤3 days"),
    ("due_10", "Due ≤10 days"),
    ("on_track", "On track (>10d)"),
    ("no_due_date", "No due date yet"),
    ("responded", "Responded"),
    ("closed", "Closed"),
]
BUCKET_LABEL = dict(BUCKETS)

# ...and one chip that is a total of several: everything still open with no
# reply filed. It leads the row because it is the question the office asks
# first - how much is left to do.
GROUPS = [("to_respond", "To respond",
           ("overdue", "due_3", "due_10", "on_track", "no_due_date"))]
GROUP_MEMBERS = {key: members for key, _, members in GROUPS}

TITLE = "Position at a glance"
CAUTION = "Draft for review — verify every figure against the portal."
XLSX_MEDIA = ("application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet")


def parse_date(text) -> date | None:
    """Portal dates are strings; this is the only place they become dates."""
    if not text:
        return None
    raw = str(text).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def is_open(status) -> bool:
    """Anything the portal has not marked Closed still needs an answer."""
    return (status or "").strip().lower() != "closed"


def bucket_of(days: int | None, open_: bool, responded: bool = False) -> str:
    """Negative days means the deadline has already gone.

    A filed reply outranks every deadline: once the portal shows "View
    Response" the notice is answered, and a passed date on an answered notice
    is history, not work.
    """
    if not open_:
        return "closed"
    if responded:
        return "responded"
    if days is None:
        return "no_due_date"
    if days < 0:
        return "overdue"
    if days <= 3:
        return "due_3"
    if days <= 10:
        return "due_10"
    return "on_track"


def _item(row: dict, today: date) -> dict:
    due = parse_date(row.get("due_date"))
    days = (due - today).days if due else None
    open_ = is_open(row.get("status"))
    # 1 filed, 0 not filed, NULL the portal never said - and "never said" is
    # treated as unanswered, because the alternative is quietly dropping a
    # notice off the list of things to do.
    responded = row.get("responded")
    return {
        "ref_id": row.get("ref_id"),
        "description": row.get("description"),
        "proceeding_name": row.get("proceeding_name"),
        "notice_us": row.get("notice_us"),
        "pan": row.get("pan"),
        "assessment_year": row.get("assessment_year"),
        "issued_on": row.get("issued_on"),
        "due_date": row.get("due_date"),
        "due_date_source": row.get("due_date_source"),
        "days_left": days,
        "bucket": bucket_of(days, open_, bool(responded)),
        "responded": responded,
        "status": row.get("status"),
        "open": open_,
        "has_pdf": bool(row.get("has_pdf")),
        "has_draft": bool(row.get("has_draft")),
        "has_due_date": bool(row.get("due_date")),
    }


def build_summary(con, today: date | None = None) -> dict:
    """Everything the report needs, counted over every notice held."""
    today = today or date.today()
    register = [_item(dict(r), today) for r in db.list_notices(con)]

    counts = {key: 0 for key, _ in BUCKETS}
    for item in register:
        counts[item["bucket"]] += 1
    for key, _, members in GROUPS:
        counts[key] = sum(counts[m] for m in members)

    # What a person has to act on: the dates that have gone, the ones about to,
    # and the notices with no date at all - which are the easiest to forget.
    # Anything already answered is not here, by virtue of its bucket.
    attention = sorted(
        (i for i in register
         if i["bucket"] in ("overdue", "due_3", "no_due_date")),
        key=lambda i: (i["days_left"] is None,
                       i["days_left"] if i["days_left"] is not None else 0))

    row = db.last_run(con)
    run = {
        # The runs table counts only what was new, so "scanned" is the whole
        # register: what the tool is holding right now.
        "notices_scanned": len(register),
        "finished": row["finished"] if row else None,
        "status": row["status"] if row else None,
        "new_this_run": (row["notices_new"] or 0) if row else 0,
        "pdfs_saved": (row["pdfs_saved"] or 0) if row else 0,
        "skipped_cached": (row["skipped_cached"] or 0) if row else 0,
    }
    return {
        "title": TITLE,
        "caution": CAUTION,
        "generated_on": today.isoformat(),
        "run": run,
        "buckets": [{"key": key, "label": label, "count": counts[key]}
                    for key, label, _ in GROUPS]
                   + [{"key": key, "label": label, "count": counts[key]}
                      for key, label in BUCKETS],
        "attention": attention,
        "register": register,
    }


def filename(summary: dict) -> str:
    return f"itr-summary-{summary['generated_on']}.xlsx"


# ------------------------------------------------------------------ Excel
ATTENTION_HEADERS = ["Client / Description", "PAN", "AY", "Section",
                     "Due date", "Days left", "Responded", "PDF saved",
                     "Draft ready"]
REGISTER_HEADERS = ["Reference ID", "Client / Description", "Proceeding", "PAN",
                    "AY", "Section", "Issued on", "Due date", "Due date from",
                    "Days left", "Position", "Responded", "Proceeding status",
                    "PDF saved", "Draft ready"]

DATE_FMT = "DD-MMM-YYYY"
HEADER_BG = "1F3864"          # the old tracker's dark blue
HEADER_FG = "FFFFFF"


def _describe(item: dict) -> str:
    return item["description"] or item["proceeding_name"] or item["ref_id"] or ""


def _yes(flag) -> str:
    return "Yes" if flag else "No"


def _replied(value) -> str:
    """Three answers, not two: the portal does not always say."""
    if value is None:
        return "Unknown"
    return "Yes" if value else "No"


def _put_date(cell, text) -> None:
    """A date the portal printed goes in as a real date; anything we cannot
    read stays as the text it was, rather than becoming a wrong date."""
    parsed = parse_date(text)
    if parsed:
        cell.value = parsed
        cell.number_format = DATE_FMT
    else:
        cell.value = text or ""


def _autosize(ws, max_width: int = 46) -> None:
    from openpyxl.utils import get_column_letter
    for column in ws.columns:
        widest = 0
        for cell in column:
            text = (cell.value.strftime("%d-%b-%Y")
                    if isinstance(cell.value, date) else str(cell.value or ""))
            widest = max(widest, len(text))
        ws.column_dimensions[get_column_letter(column[0].column)].width = \
            min(widest + 2, max_width)


def _header_row(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.append(headers)
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color=HEADER_FG)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"          # the header stays put while scrolling


def build_workbook(summary: dict) -> bytes:
    """The three sheets the office reads: Summary, Attention, All notices.

    openpyxl is imported here rather than at the top so the dashboard still
    starts on a machine where the dependency has not been installed yet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # 1 - Summary: what the old first sheet said, in the same order
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "ITR notice tool — summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Run date"
    _put_date(ws["B2"], summary["generated_on"])
    ws["A3"] = "Last sync finished"
    ws["B3"] = summary["run"]["finished"] or "no sync has finished yet"
    ws["A4"] = "Notices scanned"
    ws["B4"] = summary["run"]["notices_scanned"]
    ws["A5"] = "New this run"
    ws["B5"] = summary["run"]["new_this_run"]
    ws["A6"] = summary["caution"]
    ws["A6"].font = Font(italic=True)
    ws["A8"] = summary["title"]
    ws["A8"].font = Font(bold=True)
    for label in ("A2", "A3", "A4", "A5"):
        ws[label].font = Font(bold=True)
    at = 9
    for bucket in summary["buckets"]:
        ws.cell(row=at, column=1, value=bucket["label"]).font = Font(bold=True)
        ws.cell(row=at, column=2, value=bucket["count"])
        at += 1
    _autosize(ws)

    # 2 - Attention: overdue and due within three days, nothing else
    ws = wb.create_sheet("Attention")
    _header_row(ws, ATTENTION_HEADERS)
    for item in summary["attention"]:
        ws.append([_describe(item), item["pan"] or "", item["assessment_year"] or "",
                   item["notice_us"] or "", None, item["days_left"],
                   _replied(item["responded"]),
                   _yes(item["has_pdf"]), _yes(item["has_draft"])])
        _put_date(ws.cell(row=ws.max_row, column=5), item["due_date"])
    if not summary["attention"]:
        ws.append(["Nothing overdue or critical."])
    _autosize(ws)

    # 3 - All notices: the whole register, including where each date came from
    ws = wb.create_sheet("All notices")
    _header_row(ws, REGISTER_HEADERS)
    for item in summary["register"]:
        ws.append([item["ref_id"] or "", _describe(item),
                   item["proceeding_name"] or "", item["pan"] or "",
                   item["assessment_year"] or "", item["notice_us"] or "",
                   None, None, item["due_date_source"] or "",
                   item["days_left"], BUCKET_LABEL[item["bucket"]],
                   _replied(item["responded"]),
                   item["status"] or "", _yes(item["has_pdf"]),
                   _yes(item["has_draft"])])
        _put_date(ws.cell(row=ws.max_row, column=7), item["issued_on"])
        _put_date(ws.cell(row=ws.max_row, column=8), item["due_date"])
    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
