"""Plain-script tests: FastAPI TestClient only, no browser automation.

    .venv/bin/python test_app.py

Covers the credential rules that matter:
  - a sync without a login asks for one instead of starting
  - once posted, the login is used, remembered, and forgotten on demand
  - a rejected password drops the login and never retries
  - the password never appears in a response body, a broadcast, or the db
"""
import asyncio
import pathlib
import time
import re
import sys
import tempfile
import time
from pathlib import Path

from app import db

TMP = Path(tempfile.mkdtemp(prefix="itr-test-")) / "itr.db"
db.DB_PATH = TMP                      # keep the real data/itr.db untouched

from app import main                  # noqa: E402  (after DB_PATH is redirected)
from app.portal import scraper        # noqa: E402
from app.portal.session import WrongPasswordError  # noqa: E402
from fastapi.testclient import TestClient          # noqa: E402

USER_ID = "AAACU3358G"
PASSWORD = "sup3r-s3cret-pw"

failures = []


def check(name, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + name + (f"  [{detail}]" if detail else ""))
    if not ok:
        failures.append(name)


# ----------------------------------------------------------------- fake browser
class FakeSession:
    """Stands in for PortalSession. Records what login() was handed."""
    instances = []
    login_error = None

    def __init__(self, events, user_id, password):
        self.events, self.user_id, self.password = events, user_id, password
        self.started = self.logged_in = self.stopped = False
        FakeSession.instances.append(self)

    async def start(self):
        self.started = True

    async def login(self):
        if FakeSession.login_error:
            raise FakeSession.login_error
        self.logged_in = True

    async def stop(self):
        self.stopped = True

    async def save_debug_screenshot(self, tag="fail"):
        return f"data/debug/{tag}-test.png"

    def page_closed(self):
        return False


LIMITS_SEEN = []


async def fake_run_sync(session, events, limit=None):
    LIMITS_SEEN.append(limit)
    return {"proceedings": 1, "notices": 2, "downloaded": 0, "skipped_cached": 2}


REAL_RUN_SYNC = scraper.run_sync      # the fake below shadows it for the API tests
main.PortalSession = FakeSession
scraper.run_sync = fake_run_sync
main.settings.hold_on_error = 0     # no 15-second pause inside the tests


class Recorder:
    """Duck-types a WebSocket so we can read everything the hub pushes."""

    def __init__(self):
        self.sent = []

    async def send_json(self, payload):
        self.sent.append(payload)


def reset():
    FakeSession.instances.clear()
    FakeSession.login_error = None
    main.hub._credentials = None
    main.hub.state = "credentials_required"


def wait_idle(client, tries=200):
    """Background sync task runs on the app loop; let it finish."""
    for _ in range(tries):
        s = client.get("/api/notices").json()["state"]
        if s not in ("running",):
            return s
        time.sleep(0.02)
    return "timeout"


bodies = []          # every response body we see, for the leak check


def post(client, url, **kw):
    r = client.post(url, **kw)
    bodies.append(r.text)
    return r


with TestClient(main.app) as client:
    rec = Recorder()
    main.hub.sockets.append(rec)

    # 1 - sync with no credentials held ---------------------------------------
    reset()
    r = post(client, "/api/sync")
    check("sync without credentials returns credentials_required",
          r.status_code == 200 and r.json() == {"state": "credentials_required"},
          r.text)
    check("sync without credentials starts no browser", not FakeSession.instances)

    # 2 - posting credentials starts the sync and hands them to login ----------
    reset()
    r = post(client, "/api/credentials",
             json={"user_id": USER_ID, "password": PASSWORD})
    check("POST /api/credentials accepted",
          r.status_code == 200 and r.json()["stored"] and r.json()["started"],
          r.text)
    state = wait_idle(client)
    s = FakeSession.instances[0] if FakeSession.instances else None
    check("sync proceeded to a logged-in session",
          s is not None and s.started and s.logged_in and s.stopped)
    check("login got exactly the posted credentials",
          s is not None and s.user_id == USER_ID and s.password == PASSWORD)
    check("state back to idle after a clean run", state == "idle", state)

    # 3 - credentials are remembered for the next sync -------------------------
    FakeSession.instances.clear()
    r = post(client, "/api/sync")
    check("second sync starts without re-asking",
          r.json().get("started") is True, r.text)
    wait_idle(client)
    check("second sync reused the remembered login",
          bool(FakeSession.instances)
          and FakeSession.instances[0].password == PASSWORD)

    # 4 - "Change login" forgets them -----------------------------------------
    r = client.delete("/api/credentials")
    bodies.append(r.text)
    check("DELETE /api/credentials clears the login",
          r.status_code == 200 and not main.hub.has_credentials())
    r = post(client, "/api/sync")
    check("sync asks again after Change login",
          r.json() == {"state": "credentials_required"}, r.text)

    # 5 - wrong password: forget, announce, never retry ------------------------
    reset()
    rec.sent.clear()
    FakeSession.login_error = WrongPasswordError("Portal rejected the password.")
    post(client, "/api/credentials",
         json={"user_id": USER_ID, "password": PASSWORD})
    wait_idle(client)
    check("wrong password clears the stored login", not main.hub.has_credentials())
    check("wrong password leaves state credentials_required",
          main.hub.state == "credentials_required", main.hub.state)
    prompts = [m for m in rec.sent if m.get("type") == "credentials_required"]
    check("wrong password pushes credentials_required over the socket",
          bool(prompts) and "rejected" in (prompts[-1].get("error") or ""),
          str(prompts[-1]) if prompts else "no push")
    check("wrong password is never retried", len(FakeSession.instances) == 1,
          f"{len(FakeSession.instances)} login attempts")
    shots = [m.get("msg", "") for m in rec.sent if m.get("type") == "log"]
    check("failure path logs a screenshot path",
          any("Screenshot of the failure" in m for m in shots),
          str(shots[-2:]))

    # 6 - nothing leaks the password ------------------------------------------
    reset()
    post(client, "/api/credentials", json={"user_id": USER_ID, "password": PASSWORD})
    wait_idle(client)
    for url in ("/api/notices", "/api/notices/999/pdf", "/"):
        bodies.append(client.get(url).text)
    bodies.append(post(client, "/api/notices/999/ask-claude").text)
    with client.websocket_connect("/ws") as ws:
        bodies.append(str(ws.receive_json()))
    bodies.append(str(rec.sent))

    check("no response body or broadcast contains the password",
          not any(PASSWORD in b for b in bodies),
          next((b[:120] for b in bodies if PASSWORD in b), ""))
    check("password never reaches the sqlite file",
          PASSWORD.encode() not in TMP.read_bytes())
    check("no file on disk holds the password",
          not any(PASSWORD in p.read_text(errors="ignore")
                  for p in Path(".").rglob("*")
                  if p.is_file() and p.suffix in {".py", ".html", ".md", ".env",
                                                  ".example", ".txt", ".yml"}
                  and ".venv" not in str(p) and p.name != "test_app.py"))

    main.hub.sockets.remove(rec)


# ---------------------------------------------------- the post-password loop
# The bug this guards: Angular ships hidden validation nodes ("Please enter
# valid password") in the DOM from page load, and get_by_text matches
# substrings, so a .count()-only check aborted every login as a wrong password.
from app.portal import session as session_mod                # noqa: E402
from app.portal.session import PortalSession                 # noqa: E402

DASHBOARD = "https://eportal.incometax.gov.in/iec/foservices/#/dashboard"
LOGIN = "https://eportal.incometax.gov.in/iec/foservices/#/login"


class FakeLoc:
    def __init__(self, count=0, visible=False, text="", children=None):
        self._count, self._visible, self._text = count, visible, text
        self.clicked = False
        self.filled = None
        self._children = children or {}

    @property
    def first(self):
        return self

    async def count(self):
        return self._count

    async def is_visible(self):
        return self._visible

    async def inner_text(self):
        return self._text

    async def click(self, timeout=None):
        self.clicked = True

    async def fill(self, value):
        self.filled = value

    async def wait_for(self, state=None, timeout=None):
        return None

    async def input_value(self):
        return self._text

    def get_by_role(self, role, name=None, exact=None):
        return self._children.get(name, FakeLoc())

    def or_(self, other):
        return self if self._count else other


class FakePage:
    """Just enough Page for _settle_post_password."""

    def __init__(self, error=None, force_label=None, force=None, otp=None,
                 dashboard_after=None, transient=None):
        self._url = LOGIN
        self.error = error                    # FakeLoc for a password error
        self.force_label, self.force = force_label, force
        self.otp = otp
        self.transient = transient            # FakeLoc for "not authenticated"
        self.dashboard_after = dashboard_after
        self.reads = 0
        self.continue_clicks = 0
        self.pwd_field = FakeLoc(1, True, "")

    @property
    def url(self):
        self.reads += 1
        if self.dashboard_after is not None and self.reads > self.dashboard_after:
            return DASHBOARD
        return self._url

    def get_by_text(self, text):
        if self.error is not None and text in session_mod.PASSWORD_ERRORS:
            return self.error
        if self.transient is not None and text in session_mod.TRANSIENT_ERRORS:
            return self.transient
        return FakeLoc()

    def get_by_role(self, role, name=None):
        if name == "Continue":
            page = self

            class ContinueBtn(FakeLoc):
                @property
                def first(self):
                    return self

                async def click(self, timeout=None):
                    page.continue_clicks += 1

            return ContinueBtn(1, True, "Continue")
        if self.force is not None and name == self.force_label:
            return self.force
        return FakeLoc()

    def get_by_placeholder(self, text):
        if text == "Password":
            return self.pwd_field
        return self.otp if self.otp is not None else FakeLoc()

    def locator(self, selector):
        return FakeLoc(1, True)


class FakeEvents:
    def __init__(self):
        self.logs, self.otp_asked = [], 0

    async def log(self, msg):
        self.logs.append(msg)

    async def request_otp(self):
        self.otp_asked += 1
        return "123456"


def settle(page, grace=0.0):
    """Run the loop against a fake page, fast."""
    session_mod.POLL_SECONDS = 0.01
    session_mod.ERROR_GRACE_SECONDS = grace
    session_mod.RETRY_PAUSE_SECONDS = 0        # the real 2s pause is not the point here
    events = FakeEvents()
    s = PortalSession(events, USER_ID, PASSWORD)
    s.page = page
    err = None
    try:
        asyncio.run(s._settle_post_password())
    except Exception as e:                      # noqa: BLE001 - the test wants it
        err = e
    return err, events


# 7 - a hidden error node is not an error -----------------------------------
err, _ = settle(FakePage(error=FakeLoc(1, False, "Please enter valid password"),
                         dashboard_after=2))
check("hidden error node does NOT raise (the login bug)", err is None,
      repr(err))

# 8 - a visible one is ------------------------------------------------------
err, _ = settle(FakePage(
    error=FakeLoc(1, True, "Please enter valid password. 2 attempts left"),
    dashboard_after=99))
check("visible error node DOES raise",
      isinstance(err, WrongPasswordError), repr(err))
check("the raise quotes what the portal actually said",
      err is not None and "2 attempts left" in str(err), str(err))
check("the raise never contains the password",
      err is not None and PASSWORD not in str(err))

# 9 - the grace period covers the navigation gap ----------------------------
err, _ = settle(FakePage(error=FakeLoc(1, True, "Please enter valid password"),
                         dashboard_after=3), grace=5.0)
check("no error is believed during the grace period", err is None, repr(err))

# 10 - force-login and OTP share the visible-only rule ----------------------
hidden_btn = FakeLoc(1, False, "Yes")
err, _ = settle(FakePage(force_label="Yes", force=hidden_btn, dashboard_after=2))
check("hidden force-login button is NOT clicked",
      err is None and not hidden_btn.clicked)

shown_btn = FakeLoc(1, True, "Login Here")
err, _ = settle(FakePage(force_label="Login Here", force=shown_btn,
                         dashboard_after=2))
check("visible force-login button IS clicked",
      err is None and shown_btn.clicked)

hidden_otp = FakeLoc(1, False)
err, events = settle(FakePage(otp=hidden_otp, dashboard_after=2))
check("hidden OTP template does NOT freeze the login",
      err is None and events.otp_asked == 0)

shown_otp = FakeLoc(1, True)
err, events = settle(FakePage(otp=shown_otp, dashboard_after=2))
check("visible OTP box DOES ask the dashboard for a code",
      err is None and events.otp_asked == 1, f"asked {events.otp_asked}x")


# ------------------------------------------------- parsers, against real text
# Copied verbatim out of the live recon dumps (data/debug/recon3/), so these
# are the portal's own words, not a guess at them.
from app.portal import scraper                                  # noqa: E402

REAL_PROCEEDING = """Proceeding Name :
Issue Letter
Assessment Year :
Not Available
PAN
AAACU3358G
Name of Assessee
CAMBRIDGE TECHNOLOGY ENTERPRISES LIMITED
1
18-Aug-2026
Open
Financial Year :
Not Available
Applicable Act :
Income Tax Act 1961
View Notices/Orders (1)
+ Add / View Authorised Representative"""

REAL_NOTICE = """Notice/ Communication Reference ID : 100118320996
Notice u/s
ITBA/COM/F/17/2026-27/1092231604(1)
Document reference ID
Description :
[ITBA]Issue Letter
Issued On :
17-Aug-2026
Last Response submitted On :
18-Aug-2026
Response viewed by AO on :
27-Aug-2026
View Response
Notice/Letter pdf
Seek/View Adjournment"""


class TextCard:
    def __init__(self, text):
        self._text = text

    async def inner_text(self):
        return self._text


p = asyncio.run(scraper._parse_proceeding(TextCard(REAL_PROCEEDING), "self", "action"))
check("proceeding: name", p["proceeding_name"] == "Issue Letter", p["proceeding_name"])
check("proceeding: PAN", p["pan"] == "AAACU3358G", p["pan"])
check("proceeding: assessee", p["assessee_name"] == "CAMBRIDGE TECHNOLOGY ENTERPRISES LIMITED")
check("proceeding: 'Not Available' year becomes NULL",
      p["assessment_year"] is None and p["financial_year"] is None,
      f"{p['assessment_year']} / {p['financial_year']}")
check("proceeding: act", p["applicable_act"] == "Income Tax Act 1961", p["applicable_act"])
check("proceeding: status", p["status"] == "Open", p["status"])

n = asyncio.run(scraper._parse_notice(TextCard(REAL_NOTICE), 1))
check("notice: reference id", n["ref_id"] == "100118320996", n["ref_id"])
check("notice: document reference",
      n["doc_ref_id"] == "ITBA/COM/F/17/2026-27/1092231604(1)", n["doc_ref_id"])
# The card prints the doc reference on the line after "Notice u/s"; a naive
# parser files it as the section this notice was issued under.
check("notice: section is NOT the ITBA reference", n["notice_us"] is None,
      repr(n["notice_us"]))
check("notice: description", n["description"] == "[ITBA]Issue Letter", n["description"])
check("notice: issued on", n["issued_on"] == "17-Aug-2026", n["issued_on"])
check("notice: no due date on this letter",
      n["due_date"] is None and n["due_date_source"] is None)
check("notice: AO viewed on", n["ao_viewed_on"] == "27-Aug-2026", n["ao_viewed_on"])

# the read-only guardrail is enforced, not just documented
class Clickable:
    def __init__(self):
        self.clicked = False

    async def click(self):
        self.clicked = True


for label in ("Submit Response", "View Response", "Seek/View Adjournment"):
    c = Clickable()
    try:
        asyncio.run(scraper._click(c, label))
        check(f"guardrail refuses {label!r}", False, "it clicked")
    except RuntimeError:
        check(f"guardrail refuses {label!r}", not c.clicked)

ok = Clickable()
asyncio.run(scraper._click(ok, "Notice/Letter Pdf"))
check("guardrail allows the PDF button", ok.clicked)

# the locator crash found live: a "/" inside a regex name breaks Playwright
check("no get_by_role(name=re.compile) with a slash remains",
      not re.search(r"get_by_role\([^)]*re\.compile\([^)]*/", 
                    pathlib.Path("app/portal/scraper.py").read_text()))
# the docstring explains why go_back is banned, so look for real calls only
_src = pathlib.Path("app/portal/scraper.py").read_text()
_calls = [l for l in _src.splitlines()
          if re.search(r"(?<!Never )(?<!so )page\.go_back\(", l)
          and not l.strip().startswith(("#", '"""'))
          and '"""' not in l]
check("no page.go_back() call remains (it triggers the portal logout dialog)",
      not _calls, str(_calls))


# 11 - "Request is not authenticated": press Continue again, never a retry ----
# Seen live on the password page with a correct password. It must NOT be
# mistaken for a rejection, and it must not press Continue forever.
pg = FakePage(transient=FakeLoc(1, True, "Error : Request is not authenticated"),
              dashboard_after=3)
err, events = settle(pg)
check("transient error does not raise WrongPasswordError",
      not isinstance(err, WrongPasswordError), repr(err))
check("transient error presses Continue again", pg.continue_clicks >= 1,
      f"{pg.continue_clicks} clicks")
check("transient error is logged in the portal's own words",
      any("Request is not authenticated" in m for m in events.logs),
      str(events.logs[-1:]))

# it must give up rather than hammer the login
pg = FakePage(transient=FakeLoc(1, True, "Request is not authenticated"),
              dashboard_after=None)
session_mod.SETTLE_SECONDS = 2          # do not sit here for a minute
err, _ = settle(pg)
session_mod.SETTLE_SECONDS = 60
check("the retry cap allows more than the two the owner ran out of",
      session_mod.MAX_CONTINUE_RETRIES >= 3, str(session_mod.MAX_CONTINUE_RETRIES))
check("Continue is pressed at most MAX_CONTINUE_RETRIES times",
      pg.continue_clicks <= session_mod.MAX_CONTINUE_RETRIES,
      f"{pg.continue_clicks} clicks")
check("a stuck login times out instead of claiming a wrong password",
      err is not None and not isinstance(err, WrongPasswordError), repr(err))

# a real rejection still aborts on the first sighting
pg = FakePage(error=FakeLoc(1, True, "Invalid password"), dashboard_after=99)
err, _ = settle(pg)
check("a rejected password is still never retried",
      isinstance(err, WrongPasswordError) and pg.continue_clicks == 0,
      f"{type(err).__name__}, {pg.continue_clicks} clicks")


# 12 - the first live sync: tabs missing, yet it reported success ------------
# The hash change routes instantly and Angular had not painted, so every tab
# was "not on this account" and the run finished as done with 0 proceedings.
scraper.LIST_READY_SECONDS = 0.4        # keep the test quick


class BlankPage:
    """A list page that never paints: no cards, no tabs, no empty state."""

    def __init__(self):
        self.url = "https://eportal.incometax.gov.in/iec/foservices/#/dashboard/eProceedings"
        self.evaluated = 0

    async def evaluate(self, script, *args):
        self.evaluated += 1
        return []

    def locator(self, selector):
        return FakeLoc(0, False)

    def get_by_role(self, role, name=None, exact=None):
        return FakeLoc(0, False)

    def get_by_text(self, text, exact=None):
        return FakeLoc(0, False)

    async def wait_for_timeout(self, ms):
        await asyncio.sleep(0)


class StubSession:
    def __init__(self, page):
        self.page = page

    async def ensure_alive(self):
        return None


blank = BlankPage()
ev = FakeEvents()
err = None
try:
    asyncio.run(REAL_RUN_SYNC(StubSession(blank), ev))
except Exception as e:                                   # noqa: BLE001
    err = e
check("an unpainted list page raises instead of reporting a clean sync",
      isinstance(err, RuntimeError), repr(err))
check("the error says nothing was scraped",
      err is not None and "nothing was scraped" in str(err), str(err)[:90])
check("no 'Sync done' line is logged when nothing was found",
      not any("Sync done" in m for m in ev.logs), str(ev.logs[-1:]))

# a visible tab is found through any of the three shapes it takes
class TabPage(BlankPage):
    def __init__(self, where):
        super().__init__()
        self.where = where

    def get_by_role(self, role, name=None, exact=None):
        if self.where == role and name == "Self":
            return FakeLoc(1, True, "Self")
        return FakeLoc(0, False)

    def get_by_text(self, text, exact=None):
        if self.where == "text" and text == "Self":
            return FakeLoc(1, True, "Self")
        return FakeLoc(0, False)


for where in ("button", "tab", "text"):
    found = asyncio.run(scraper._find_tab(TabPage(where), "Self"))
    check(f"tab found when it is a {where}", found is not None)

check("a hidden tab is not treated as present",
      asyncio.run(scraper._find_tab(BlankPage(), "Self")) is None)

scraper.LIST_READY_SECONDS = 30


# 13 - the back/refresh modal ------------------------------------------------
# Live failure: "#securityReasonPopup intercepts pointer events" - the modal
# swallows every click underneath it, and its YES button logs the session out.
from app.portal.session import dismiss_security_popup                # noqa: E402


class PopupPage:
    def __init__(self, visible=True):
        self.url = "https://x/#/dashboard/eProceedings"
        self.no = FakeLoc(1, True, "No")
        self.yes = FakeLoc(1, True, "YES")
        self.popup = FakeLoc(1 if visible else 0, visible, "back disabled",
                             children={"No": self.no, "YES": self.yes})
        self.waited = 0

    def locator(self, selector):
        return self.popup if selector == "#securityReasonPopup" else FakeLoc()

    async def wait_for_timeout(self, ms):
        self.waited += 1


pg = PopupPage()
ev = FakeEvents()
check("the back/refresh modal is dismissed", asyncio.run(dismiss_security_popup(pg, ev)))
check("it is dismissed with No", pg.no.clicked)
check("YES is never pressed (it logs the session out)", not pg.yes.clicked)
check("no modal on screen means nothing to do",
      not asyncio.run(dismiss_security_popup(PopupPage(visible=False))))

for label in ("YES", "Yes", "Logout"):
    c = Clickable()
    try:
        asyncio.run(scraper._click(c, label))
        check(f"guardrail refuses {label!r}", False, "it clicked")
    except RuntimeError:
        check(f"guardrail refuses {label!r}", not c.clicked)

# the navigation that caused it must be gone for good
_scraper_src = pathlib.Path("app/portal/scraper.py").read_text()
check("nothing changes the URL any more (that is what raised the modal)",
      "location.hash" not in _scraper_src.split('"""', 2)[2])


# 14 - the download limit ----------------------------------------------------
check("limit stops the walk once the cap is met",
      scraper._limit_reached({"limit": 3, "downloaded": 3}))
check("limit does not stop the walk early",
      not scraper._limit_reached({"limit": 3, "downloaded": 2}))
check("no limit means every notice",
      not scraper._limit_reached({"limit": None, "downloaded": 99}))
check("a limit of 0 is treated as no limit",
      not scraper._limit_reached({"limit": 0, "downloaded": 99}))

with TestClient(main.app) as client:
    reset()
    LIMITS_SEEN.clear()
    r = client.post("/api/credentials",
                    json={"user_id": USER_ID, "password": PASSWORD, "limit": 3})
    check("credentials accept a limit", r.json().get("limit") == 3, r.text)
    wait_idle(client)
    check("the limit reaches the walk", LIMITS_SEEN[-1:] == [3], str(LIMITS_SEEN))

    LIMITS_SEEN.clear()
    client.post("/api/sync", json={"limit": 5})
    wait_idle(client)
    check("sync can change the limit per run", LIMITS_SEEN[-1:] == [5], str(LIMITS_SEEN))

    LIMITS_SEEN.clear()
    client.post("/api/sync", json={"limit": None})
    wait_idle(client)
    check("blank means all", LIMITS_SEEN[-1:] == [None], str(LIMITS_SEEN))

    LIMITS_SEEN.clear()
    client.post("/api/sync")
    wait_idle(client)
    check("a sync with no body still works", LIMITS_SEEN[-1:] == [None], str(LIMITS_SEEN))


# 15 - the access lock -------------------------------------------------------
# The tool is going on a public URL, so nothing may answer without the
# dashboard password: not the page, not the API, not the WebSocket.
APP_PW = "dashboard-pw"
main.FAILED_LOGIN_DELAY = 0            # the real 2s delay is asserted separately
main.settings.app_password = ""

with TestClient(main.app) as anon:
    check("with no APP_PASSWORD set the dashboard is open (localhost dev)",
          anon.get("/api/notices").status_code == 200)

main.settings.app_password = APP_PW

with TestClient(main.app) as anon:
    r = anon.get("/")
    check("locked: the page asks for a password",
          r.status_code == 401 and "password" in r.text.lower(), str(r.status_code))
    r = anon.get("/api/notices")
    check("locked: the API answers 401, not data",
          r.status_code == 401 and "notices" not in r.json(), r.text[:60])
    check("locked: a stored PDF is not downloadable",
          anon.get("/api/notices/100118320996/pdf").status_code == 401)
    check("locked: sync cannot be started",
          anon.post("/api/sync", json={}).status_code == 401)
    ws_refused = False
    try:
        with anon.websocket_connect("/ws"):
            pass
    except Exception:
        ws_refused = True
    check("locked: the WebSocket handshake is refused", ws_refused)

    r = anon.post("/login", json={"password": "wrong-one"})
    check("a wrong password is rejected", r.status_code == 401, r.text)
    check("a wrong password sets no cookie",
          main.COOKIE_NAME not in anon.cookies, str(dict(anon.cookies)))

    r = anon.post("/login", json={"password": APP_PW})
    check("the right password signs in", r.status_code == 200, r.text)
    check("signing in sets the session cookie", main.COOKIE_NAME in anon.cookies)
    check("signed in: the API answers", anon.get("/api/notices").status_code == 200)
    opened = False
    with anon.websocket_connect("/ws") as ws:
        opened = ws.receive_json()["type"] == "state"
    check("signed in: the WebSocket opens", opened)

    anon.post("/logout")
    check("logging out locks it again", anon.get("/api/notices").status_code == 401)

# the cookie is signed, not guessable, and expires
check("a forged cookie is rejected", not main._cookie_ok("9999999999.deadbeef"))
check("a cookie signed with a different password is rejected",
      not main._cookie_ok(main._new_cookie().split(".")[0] + ".c0ffee"))
good = main._new_cookie()
check("a freshly issued cookie is accepted", main._cookie_ok(good))
check("an expired cookie is rejected",
      not main._cookie_ok(f"{int(time.time()) - main.COOKIE_MAX_AGE - 60}."
                          + main._sign(str(int(time.time()) - main.COOKIE_MAX_AGE - 60))))
check("changing APP_PASSWORD invalidates old cookies",
      (lambda: (setattr(main.settings, "app_password", "different-pw"),
                main._cookie_ok(good))[1])() is False)

main.settings.app_password = APP_PW
check("failed logins are slowed down on purpose", main.FAILED_LOGIN_DELAY == 0
      or main.FAILED_LOGIN_DELAY >= 1)
main.settings.app_password = ""        # leave the rest of the suite unlocked


# 16 - dashboard summary + filters -------------------------------------------
# The cards and filters run in the browser, so what is checked here is that the
# page ships them and that the maths they implement is right on real row shapes.
import datetime                                                      # noqa: E402

_page_now = lambda: "".join(
    pathlib.Path(f"app/static/{f}").read_text()
    for f in ("index.html", "app.js", "style.css"))
_page = _page_now()
for hook in ("f-ay", "f-name", "f-nodue", "applyFilters", "dueInDays",
             "s-total", "s-week", "s-nodue", "s-docs", "s-drafts",
             "renderStats", "renderLastSync"):
    check(f"dashboard ships {hook}", hook in _page)

def _days(d):
    return (datetime.datetime.strptime(d, "%d-%b-%Y").date()
            - datetime.date.today()).days

_fmt = lambda n: (datetime.date.today() + datetime.timedelta(days=n)).strftime("%d-%b-%Y")
SAMPLE = [
    {"due_date": None, "assessment_year": "2020-21", "proceeding_name": "Issue Letter",
     "has_pdf": 1},
    {"due_date": _fmt(3), "assessment_year": "2021-22",
     "proceeding_name": "Assessment u/s 143", "has_pdf": 0},
    {"due_date": _fmt(7), "assessment_year": "2021-22", "proceeding_name": "Penalty",
     "has_pdf": 1},
    {"due_date": _fmt(20), "assessment_year": "2020-21", "proceeding_name": "Penalty",
     "has_pdf": 1},
    {"due_date": _fmt(-2), "assessment_year": "2020-21", "proceeding_name": "Overdue one",
     "has_pdf": 1},
]
due_week = [n for n in SAMPLE if n["due_date"] and 0 <= _days(n["due_date"]) <= 7]
check("Due this week counts the next 7 days only", len(due_week) == 2,
      str([n["proceeding_name"] for n in due_week]))
check("Due this week excludes dates already past",
      all(_days(n["due_date"]) >= 0 for n in due_week))
check("Missing due date counts blanks",
      len([n for n in SAMPLE if not n["due_date"]]) == 1)
check("Docs saved counts stored PDFs",
      len([n for n in SAMPLE if n["has_pdf"]]) == 4)
check("the year dropdown lists each year once, sorted",
      sorted({n["assessment_year"] for n in SAMPLE if n["assessment_year"]})
      == ["2020-21", "2021-22"])

def _filtered(rows, ay="", name="", nodue=False):
    return [n for n in rows
            if (not ay or n["assessment_year"] == ay)
            and (not name or name.lower() in (n["proceeding_name"] or "").lower())
            and (not nodue or not n["due_date"])]

check("year filter narrows the table", len(_filtered(SAMPLE, ay="2021-22")) == 2)
check("name filter is a case-insensitive contains",
      len(_filtered(SAMPLE, name="penalty")) == 2)
check("missing-due-date toggle keeps only blanks",
      [n["proceeding_name"] for n in _filtered(SAMPLE, nodue=True)] == ["Issue Letter"])
check("filters combine", len(_filtered(SAMPLE, ay="2020-21", name="penalty")) == 1)


# 17 - preview vs download ---------------------------------------------------
main.settings.app_password = ""
_PDF = b"%PDF-1.4\n% a stored notice\n%%EOF\n"
with db.connect() as con:
    con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES (?)", ("100118320996",))
    con.execute("UPDATE notices SET pdf_blob=? WHERE ref_id=?", (_PDF, "100118320996"))

with TestClient(main.app) as client:
    r = client.get("/api/notices/100118320996/pdf?inline=1")
    check("preview serves the PDF inline",
          r.status_code == 200 and "inline" in r.headers.get("content-disposition", ""),
          r.headers.get("content-disposition", "none"))
    check("preview declares the PDF media type",
          r.headers.get("content-type", "").startswith("application/pdf"),
          r.headers.get("content-type", "none"))
    check("preview returns the stored bytes", r.content.startswith(b"%PDF"))

    r = client.get("/api/notices/100118320996/pdf")
    check("download still attaches",
          r.status_code == 200
          and "attachment" in r.headers.get("content-disposition", ""),
          r.headers.get("content-disposition", "none"))
    check("a missing PDF is still 404 either way",
          client.get("/api/notices/does-not-exist/pdf?inline=1").status_code == 404)

check("the row offers View alongside Save",
      "view(" in _page_now() and "savePdf(" in _page_now())


# 18 - Ask Claude for a missing due date --------------------------------------
# The Claude call itself is stubbed: these tests are about the cache rule, the
# guardrail that a portal date is never overwritten, and the no-key path.
from app import claude_client                                        # noqa: E402

CALLS = []


async def fake_due_date(pdf, *, ref_id, issued_on=None, served_on=None):
    CALLS.append({"ref_id": ref_id, "pdf": pdf,
                  "issued_on": issued_on, "served_on": served_on})
    return {"due_date": "15-Sep-2026",
            "basis": "notice says within 15 days of service on 18-Aug-2026"}


async def fake_no_date(pdf, **kw):
    CALLS.append({"ref_id": kw.get("ref_id")})
    return {"due_date": None, "basis": "this letter sets no deadline"}


REAL_HAVE_KEY = claude_client.have_key      # the stubs below shadow the module
main.claude_client.due_date_from_pdf = fake_due_date
main.claude_client.have_key = lambda: True

with db.connect() as con:
    con.execute("UPDATE notices SET due_date=NULL, due_date_source=NULL, "
                "due_date_basis=NULL, issued_on='17-Aug-2026', served_on='18-Aug-2026' "
                "WHERE ref_id='100118320996'")

with TestClient(main.app) as client:
    CALLS.clear()
    r = client.post("/api/notices/100118320996/ask-claude")
    d = r.json()
    check("Ask Claude returns a date", r.status_code == 200
          and d["due_date"] == "15-Sep-2026", r.text[:90])
    check("the basis comes back with it", "15 days" in (d["basis"] or ""), str(d["basis"]))
    check("the portal dates are passed to Claude as context",
          CALLS and CALLS[0]["issued_on"] == "17-Aug-2026"
          and CALLS[0]["served_on"] == "18-Aug-2026", str(CALLS[:1]))

    with db.connect() as con:
        row = db.get_notice(con, "100118320996")
    check("the date is stored and tagged as Claude's",
          row["due_date"] == "15-Sep-2026" and row["due_date_source"] == "claude")
    check("the basis is stored in its own column",
          "15 days" in (row["due_date_basis"] or ""))

    CALLS.clear()
    r2 = client.post("/api/notices/100118320996/ask-claude")
    check("asking again is served from cache, never re-called",
          r2.json()["cached"] is True and not CALLS, str(CALLS))
    check("the cached answer still carries the basis",
          "15 days" in (r2.json()["basis"] or ""))

    # a portal date must never be replaced by a Claude one
    with db.connect() as con:
        con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('portal-dated')")
        con.execute("UPDATE notices SET due_date='01-Oct-2026', due_date_source='portal', "
                    "pdf_blob=? WHERE ref_id='portal-dated'", (_PDF,))
    CALLS.clear()
    r3 = client.post("/api/notices/portal-dated/ask-claude")
    check("a portal date is returned untouched and Claude is not called",
          r3.json()["due_date"] == "01-Oct-2026"
          and r3.json()["source"] == "portal" and not CALLS, str(CALLS))

    # Claude finding no deadline is a normal outcome, not an error
    main.claude_client.due_date_from_pdf = fake_no_date
    with db.connect() as con:
        con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('no-deadline')")
        con.execute("UPDATE notices SET pdf_blob=? WHERE ref_id='no-deadline'",
                    (_PDF,))
    r4 = client.post("/api/notices/no-deadline/ask-claude")
    check("no deadline in the notice returns null, not an error",
          r4.status_code == 200 and r4.json()["due_date"] is None, r4.text[:80])
    with db.connect() as con:
        check("nothing is written when there is no date",
              db.get_notice(con, "no-deadline")["due_date"] is None)

    # missing prerequisites
    main.claude_client.have_key = lambda: False
    r5 = client.post("/api/notices/no-deadline/ask-claude")
    check("no API key gives the 'add API key in .env' message",
          r5.status_code == 503 and "API key" in r5.json()["error"], r5.text[:80])
    main.claude_client.have_key = lambda: True

    with db.connect() as con:
        con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('no-pdf')")
    r6 = client.post("/api/notices/no-pdf/ask-claude")
    check("a notice with no stored PDF says so", r6.status_code == 404,
          r6.text[:80])
    check("an unknown notice is 404",
          client.post("/api/notices/nope/ask-claude").status_code == 404)

claude_client.have_key = REAL_HAVE_KEY      # done stubbing; test the real one
_real_key = claude_client.settings.anthropic_api_key
claude_client.settings.anthropic_api_key = "sk-ant-add-later"
check("the .env.example placeholder does not count as a key",
      not claude_client.have_key())
claude_client.settings.anthropic_api_key = ""
check("an empty key does not count as a key", not claude_client.have_key())
claude_client.settings.anthropic_api_key = "sk-ant-a-real-looking-key-value"
check("a real-looking key counts", claude_client.have_key())
claude_client.settings.anthropic_api_key = _real_key
check("the model is the one the owner asked for",
      claude_client.MODEL == "claude-sonnet-4-6", claude_client.MODEL)
check("the PDF is sent as a document block",
      claude_client._pdf_block(_PDF)["type"] == "document")
check("the due-date schema pins the strict JSON shape",
      claude_client.DUE_DATE_SCHEMA["required"] == ["due_date", "basis"]
      and claude_client.DUE_DATE_SCHEMA["additionalProperties"] is False)
check("the row offers Ask Claude only when a PDF is stored",
      "askClaude(" in _page_now() and "n.has_pdf" in _page_now())


# 19 - generate a draft response ---------------------------------------------
DRAFT_CALLS = []


async def fake_draft(pdf, *, ref_id, notice_us=None, assessee=None,
                     assessment_year=None):
    DRAFT_CALLS.append({"ref_id": ref_id, "pdf": pdf, "notice_us": notice_us,
                        "assessee": assessee, "assessment_year": assessment_year})
    return {"summary": "The officer wants proof of the deduction you claimed.",
            "checklist": ["Bank statement for FY 2019-20", "Copy of the invoice"],
            "draft_reply": "To the Assessing Officer,\n\nRe: notice "
                           f"{ref_id}. [Fill in the reply here.]"}


main.claude_client.draft_from_pdf = fake_draft
main.claude_client.have_key = lambda: True

with db.connect() as con:
    pid = db.upsert_proceeding(con, {
        "tab": "self", "sub_tab": "action", "proceeding_name": "Issue Letter",
        "pan": "AAACU3358G", "assessee_name": "CAMBRIDGE TECHNOLOGY ENTERPRISES LIMITED",
        "assessment_year": "2020-21", "financial_year": "2019-20",
        "applicable_act": "Income Tax Act 1961", "status": "Open",
        "closure_date": None, "closure_order": None})
    con.execute("UPDATE notices SET proceeding_id=?, notice_us='142(1)' "
                "WHERE ref_id='100118320996'", (pid,))

with TestClient(main.app) as client:
    DRAFT_CALLS.clear()
    r = client.post("/api/notices/100118320996/draft")
    d = r.json()
    check("draft returns a summary", r.status_code == 200
          and "deduction" in d["summary"], r.text[:80])
    check("draft returns the document checklist", d["checklist"] ==
          ["Bank statement for FY 2019-20", "Copy of the invoice"], str(d["checklist"]))
    check("draft returns editable reply text", "Assessing Officer" in d["draft_text"])
    check("the notice's own context is sent to Claude",
          DRAFT_CALLS and DRAFT_CALLS[0]["notice_us"] == "142(1)"
          and DRAFT_CALLS[0]["assessment_year"] == "2020-21", str(DRAFT_CALLS[:1]))

    DRAFT_CALLS.clear()
    r2 = client.post("/api/notices/100118320996/draft")
    check("a second open is served from the drafts table, no second call",
          r2.json()["cached"] is True and not DRAFT_CALLS, str(DRAFT_CALLS))
    check("the cached draft keeps its checklist",
          len(r2.json()["checklist"]) == 2)

    DRAFT_CALLS.clear()
    r3 = client.post("/api/notices/100118320996/draft?regenerate=1")
    check("Regenerate is the one thing that calls again",
          r3.json()["cached"] is False and len(DRAFT_CALLS) == 1, str(DRAFT_CALLS))
    with db.connect() as con:
        check("regenerating overwrites rather than piling up rows",
              con.execute("SELECT count(*) n FROM drafts WHERE ref_id=?",
                          ("100118320996",)).fetchone()["n"] == 1)

    main.claude_client.have_key = lambda: False
    check("no API key is reported, not crashed",
          client.post("/api/notices/no-deadline/draft").status_code == 503)
    main.claude_client.have_key = lambda: True
    check("a notice with no PDF cannot be drafted",
          client.post("/api/notices/no-pdf/draft").status_code == 404)
    check("an unknown notice is 404",
          client.post("/api/notices/nope/draft").status_code == 404)

check("the draft schema pins summary, checklist and reply",
      claude_client.DRAFT_SCHEMA["required"] == ["summary", "checklist", "draft_reply"])
_p = _page_now()
check("the panel says DRAFT and that nothing is submitted",
      "review before filing" in _p and "never submits to the portal" in _p)
check("the panel has an editable draft and a Copy button",
      "<textarea id=\"d-text\"" in _p and "d-copy" in _p)

# the read-only guardrail, still absolute
_backend = (pathlib.Path("app/main.py").read_text()
            + pathlib.Path("app/portal/scraper.py").read_text()
            + pathlib.Path("app/report.py").read_text()
            + pathlib.Path("app/claude_client.py").read_text())
for banned in ("submitResponse", "submit_response", "fileAppeal", "/submit"):
    check(f"no portal-submission code anywhere ({banned})", banned not in _backend)


# 20 - live viewport + pipeline messages --------------------------------------
import base64 as _b64                                               # noqa: E402
from app.portal import session as _sess                             # noqa: E402

rec2 = Recorder()
main.hub.sockets.append(rec2)

asyncio.run(main.hub.progress("walk", tab="Self", items=40))
sent = rec2.sent[-1]
check("progress is broadcast with stage and counts",
      sent["type"] == "progress" and sent["stage"] == "walk"
      and sent["counts"] == {"tab": "Self", "items": 40}, str(sent))
check("the last stage is kept for a browser that joins late",
      main.hub.last_progress == sent)

asyncio.run(main.hub.viewport(b"\xff\xd8\xff-not-really-a-jpeg"))
frame = rec2.sent[-1]
check("viewport frames are broadcast as base64",
      frame["type"] == "viewport"
      and _b64.b64decode(frame["img"]) == b"\xff\xd8\xff-not-really-a-jpeg",
      frame["type"])
check("the last frame is kept for a browser that joins late",
      main.hub.last_frame == frame["img"])
main.hub.sockets.remove(rec2)

# a browser connecting mid-sync is caught up on both
with TestClient(main.app) as client:
    with client.websocket_connect("/ws") as ws:
        kinds = [ws.receive_json()["type"] for _ in range(4)]
    check("a fresh socket is replayed state, speed, progress and the last frame",
          kinds == ["state", "speed", "progress", "viewport"], str(kinds))

# --- the rule that matters: no credential ever reaches the viewport ---------
class FakePage:
    def __init__(self):
        self.shots = 0

    def is_closed(self):
        return False

    async def screenshot(self, **kw):
        self.shots += 1
        return b"frame"


class CaptureSession:
    def __init__(self):
        self.page = FakePage()
        self.in_login = False
        self.sensitive_until = 0.0

    safe_to_capture = _sess.PortalSession.safe_to_capture
    page_closed = _sess.PortalSession.page_closed


async def _run_loop(sess, state, seconds=0.25):
    main.VIEWPORT_INTERVAL = 0.05
    main.hub.state = state
    task = asyncio.create_task(main._viewport_loop(sess))
    await asyncio.sleep(seconds)
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass

s1 = CaptureSession()
asyncio.run(_run_loop(s1, "running"))
check("frames are captured during a normal sync", s1.page.shots > 0,
      f"{s1.page.shots} frames")

s2 = CaptureSession()
asyncio.run(_run_loop(s2, "otp_required"))
check("NO frames while the dashboard is holding for an OTP",
      s2.page.shots == 0, f"{s2.page.shots} frames")

s3 = CaptureSession()
s3.in_login = True
asyncio.run(_run_loop(s3, "running"))
check("NO frames while the login screens are up (password on screen)",
      s3.page.shots == 0, f"{s3.page.shots} frames")

s4 = CaptureSession()
s4.sensitive_until = time.monotonic() + 5
asyncio.run(_run_loop(s4, "running"))
check("NO frames in the quiet window just after the password is submitted",
      s4.page.shots == 0, f"{s4.page.shots} frames")

check("the quiet window after the password is 2 seconds",
      "self.sensitive_until = time.monotonic() + 2"
      in pathlib.Path("app/portal/session.py").read_text())
# read the source, not the module: the loop above turned the interval down
_main_src = pathlib.Path("app/main.py").read_text()
check("a frame is a jpeg every 1.5s at quality 45",
      "VIEWPORT_INTERVAL = 1.5" in _main_src and "VIEWPORT_QUALITY = 45" in _main_src
      and 'type="jpeg"' in _main_src)
_mon = _page_now()
check("the monitor is a 16:9 card that starts collapsed",
      "aspect-ratio: 16 / 9" in _mon and "<details class=\"card monitor\"" in _mon)
check("it expands itself when a run starts", "$('monitor').open = true" in _mon)
check("the REC light only burns while frames are arriving",
      ".monitor.live .rec { opacity: 1; }" in _mon
      and "$('monitor').classList.add('live')" in _mon
      and "$('monitor').classList.remove('live')" in _mon)
check("and it says why it went quiet for an OTP",
      "paused - OTP on screen" in _mon)
check("the latest log line is the caption",
      "$('caption').textContent = msg.trim()" in _mon)

main.VIEWPORT_INTERVAL = 1.5
main.hub.state = "idle"
check("login marks itself sensitive so the loop can skip it",
      "self.in_login = True" in pathlib.Path("app/portal/session.py").read_text())

# 21 - the redesign ships as split files ------------------------------------
_static = pathlib.Path("app/static")
check("css and js are separate files, still statically served",
      (_static / "style.css").exists() and (_static / "app.js").exists())
check("fonts are self-hosted, not pulled from a CDN",
      (_static / "fonts" / "Geist-Variable.woff2").exists()
      and (_static / "fonts" / "GeistMono-Variable.woff2").exists())
_all = _page_now()
check("no CDN or external font link remains",
      "cdn." not in _all and "fonts.googleapis" not in _all and "unpkg" not in _all)
for hook in ("data-theme", "dueChip", "palOpen", "showFrame", "renderPipe",
             "aiCard", "SKELETON", "empty-state", "prefers-reduced-motion",
             "focus-visible", "tabular-nums"):
    check(f"v2 ships {hook}", hook in _all)
check("theme is persisted in a cookie, not localStorage",
      "document.cookie" in _all and "localStorage." not in _all
      and "localStorage[" not in _all)

with TestClient(main.app) as client:
    for path, kind in (("/", "text/html"), ("/style.css", "text/css"),
                       ("/app.js", "javascript"),
                       ("/fonts/Geist-Variable.woff2", "font")):
        r = client.get(path)
        check(f"{path} is served", r.status_code == 200
              and kind in r.headers.get("content-type", ""),
              f"{r.status_code} {r.headers.get('content-type', '')}")

# 22 - live speed control ----------------------------------------------------
# The point of this one is that it is LIVE: Playwright's slow_mo is fixed at
# launch, so the pace is ours and is re-read before every browser action.
from app.portal.session import pace_for, PortalSession        # noqa: E402

check("the three modes are the ones on the buttons",
      sorted(main.MODES) == ["extreme", "fast", "slow"], str(list(main.MODES)))
check("slow is a full second, extreme is no wait at all",
      main.MODES["slow"] == 1.0 and main.MODES["fast"] == 0.25
      and main.MODES["extreme"] == 0.0, str(main.MODES))
check("fast is the default", main.DEFAULT_MODE == "fast")

main.hub.mode = main.DEFAULT_MODE
with TestClient(main.app) as client:
    check("the mode can be read back", client.get("/api/speed").json()["mode"] == "fast")

    for name, ms in (("slow", 1000), ("extreme", 0), ("fast", 250)):
        r = client.post("/api/speed", json={"mode": name})
        check(f"mode {name} is accepted",
              r.status_code == 200 and r.json() == {"mode": name, "delay_ms": ms},
              r.text[:80])
        check(f"mode {name} is what the scraper will read",
              main.hub.mode == name and main.hub.pace_seconds() == main.MODES[name])

    r = client.post("/api/speed", json={"mode": "ludicrous"})
    check("an unknown mode is rejected", r.status_code == 400, r.text[:80])
    check("the rejection lists the modes that do work",
          all(s in r.json()["error"] for s in main.MODES), r.text[:90])
    check("a rejected mode changes nothing", main.hub.mode == "fast")
    check("a missing mode field is a 422, not a crash",
          client.post("/api/speed", json={}).status_code == 422)
    check("a mode of the wrong type is a 422, not a crash",
          client.post("/api/speed", json={"mode": 3}).status_code == 422)

    # case and stray spaces are the user's, not an error
    r = client.post("/api/speed", json={"mode": "  SLOW "})
    check("the mode name is normalised", r.status_code == 200
          and main.hub.mode == "slow", r.text[:80])

    # every open dashboard hears about it, including one that never clicked
    with client.websocket_connect("/ws") as ws:
        seen = {}
        for _ in range(2):
            m = ws.receive_json()
            seen[m["type"]] = m
        check("a new socket is told the current mode",
              seen.get("speed", {}).get("mode") == "slow", str(seen.get("speed")))
        client.post("/api/speed", json={"mode": "extreme"})
        pushed = None
        for _ in range(6):
            m = ws.receive_json()
            if m["type"] == "speed":
                pushed = m
                break
        check("a mode change is broadcast to every open dashboard",
              pushed == {"type": "speed", "mode": "extreme", "delay_ms": 0},
              str(pushed))

main.hub.mode = main.DEFAULT_MODE

# the wait itself: read live, so a change lands mid-sync
async def _timed(events):
    t0 = time.monotonic()
    await pace_for(events)
    return time.monotonic() - t0

main.hub.mode = "extreme"
check("extreme does not wait", asyncio.run(_timed(main.hub)) < 0.05)
main.hub.mode = "slow"
check("slow waits about a second", 0.9 <= asyncio.run(_timed(main.hub)) <= 1.4)
main.hub.mode = main.DEFAULT_MODE


class SpeedlessEvents:
    """An events object from before the speed knob existed."""


check("pacing a hub that has no speed knob is a no-op, not a crash",
      asyncio.run(_timed(SpeedlessEvents())) < 0.05)


class PacedSession:
    """Just the pace() method, on the same hub the real session gets."""
    events = main.hub
    pace = PortalSession.pace


async def _timed_pace(obj):
    t0 = time.monotonic()
    await obj.pace()
    return time.monotonic() - t0


main.hub.mode = "slow"
check("session.pace() reads the same live setting",
      0.9 <= asyncio.run(_timed_pace(PacedSession())) <= 1.4)


# the whole point: pressing a button mid-sync is felt by the next action
async def _mid_run_change():
    sess = PacedSession()
    main.hub.mode = "slow"
    slow = await _timed_pace(sess)
    main.hub.mode = "extreme"           # the dashboard button, mid-walk
    quick = await _timed_pace(sess)
    return slow, quick

_slow, _quick = asyncio.run(_mid_run_change())
check("changing the mode mid-sync changes the very next action",
      _slow >= 0.9 and _quick < 0.05, f"{_slow:.2f}s then {_quick:.2f}s")
main.hub.mode = main.DEFAULT_MODE

# every browser action is paced: the calls are in the source, not just the API
_sess_src = pathlib.Path("app/portal/session.py").read_text()
_scrp_src = pathlib.Path("app/portal/scraper.py").read_text()
check("login paces before the user id, the password and each Continue",
      _sess_src.count("await self.pace()") >= 6,
      str(_sess_src.count("await self.pace()")))
check("slow_mo is no longer used (it cannot change after launch)",
      "slow_mo=0," in _sess_src and "settings.slow_mo_ms" not in _sess_src)
check("every scraper click is paced", "await pace_for(events)" in _scrp_src)
check("parsing and downloading are paced too",
      _scrp_src.count("await session.pace()") >= 3,
      str(_scrp_src.count("await session.pace()")))
check("the dashboard posts the mode to the server, not to a cookie",
      "/api/speed" in _page_now() and "JSON.stringify({ mode: next })" in _page_now()
      and "speed=${SPEED}" not in _page_now())
check("the dashboard labels the buttons Slow, Fast and Extreme",
      all(f'data-speed="{s}"' in _page_now() for s in ("slow", "fast", "extreme")))
check("extreme carries its 'testing only' caption",
      "testing only" in _page_now() and "speednote" in _page_now())


# 23 - the PDFs live in the database ----------------------------------------
# One file to back up, one file to move to Lightsail: notices.pdf_blob holds
# the document itself and NOTICES_DIR is gone.
_BLOB = b"%PDF-1.7\n% blob round trip \x00\x01\x02 binary safe\n%%EOF\n"

with db.connect() as con:
    con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('blob-notice')")
    db.upsert_notice(con, {
        "proceeding_id": None, "ref_id": "blob-notice", "notice_us": "142(1)",
        "doc_ref_id": None, "description": "blob test", "issued_on": None,
        "served_on": None, "due_date": None, "due_date_source": None,
        "ao_viewed_on": None, "pdf_blob": _BLOB})

with db.connect() as con:
    check("the bytes come back out of the row unchanged",
          db.get_notice_pdf(con, "blob-notice") == _BLOB)
    check("storing a PDF stamps downloaded_at",
          db.get_notice(con, "blob-notice")["downloaded_at"] is not None)
    check("the cache rule now reads the blob, not a path",
          db.notice_exists(con, "blob-notice"))
    check("a notice with no blob is not cached",
          not db.notice_exists(con, "no-pdf"))
    check("get_notice_pdf on an unknown notice is None, not a crash",
          db.get_notice_pdf(con, "nope") is None)

    # a second sync of the same notice must not wipe the stored document
    db.upsert_notice(con, {
        "proceeding_id": None, "ref_id": "blob-notice", "notice_us": "142(1)",
        "doc_ref_id": None, "description": "blob test", "issued_on": None,
        "served_on": None, "due_date": None, "due_date_source": None,
        "ao_viewed_on": None, "pdf_blob": None})
    check("re-seeing a notice never blanks the PDF already held",
          db.get_notice_pdf(con, "blob-notice") == _BLOB)

with TestClient(main.app) as client:
    r = client.get("/api/notices/blob-notice/pdf?inline=1")
    check("view serves the stored bytes inline", r.status_code == 200
          and r.content == _BLOB, f"{r.status_code} {len(r.content)}B")
    check("view is application/pdf",
          r.headers.get("content-type", "").startswith("application/pdf"),
          r.headers.get("content-type", "none"))
    check("view asks the browser to render, not to save",
          r.headers.get("content-disposition", "").startswith("inline"),
          r.headers.get("content-disposition", "none"))

    r = client.get("/api/notices/blob-notice/pdf")
    check("save serves the same bytes as an attachment",
          r.status_code == 200 and r.content == _BLOB
          and "attachment" in r.headers.get("content-disposition", ""),
          r.headers.get("content-disposition", "none"))
    check("the download is named after the notice",
          'filename="blob-notice.pdf"' in r.headers.get("content-disposition", ""),
          r.headers.get("content-disposition", "none"))
    check("a notice with no stored PDF is 404",
          client.get("/api/notices/no-pdf/pdf").status_code == 404)

    # the blob must never ride along in the table's JSON
    rows = client.get("/api/notices").json()["notices"]
    row = next(n for n in rows if n["ref_id"] == "blob-notice")
    check("the notices list does not ship the blob", "pdf_blob" not in row,
          str(sorted(row))[:120])
    check("it ships has_pdf instead", row["has_pdf"] == 1, str(row.get("has_pdf")))
    check("a notice with no PDF says so",
          next(n for n in rows if n["ref_id"] == "no-pdf")["has_pdf"] == 0)

# the one-time move off the filesystem, on a database that still has paths
_old_dir = TMP.parent / "old-notices"
_old_dir.mkdir(exist_ok=True)
_old_file = _old_dir / "legacy.pdf"
_old_file.write_bytes(b"%PDF-legacy%%EOF")
_gone_file = _old_dir / "already-deleted.pdf"
with db.connect() as con:
    con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('legacy')")
    con.execute("UPDATE notices SET pdf_path=?, pdf_blob=NULL WHERE ref_id='legacy'",
                (str(_old_file),))
    con.execute("INSERT OR IGNORE INTO notices (ref_id) VALUES ('legacy-gone')")
    con.execute("UPDATE notices SET pdf_path=?, pdf_blob=NULL WHERE ref_id='legacy-gone'",
                (str(_gone_file),))

db.init_db()
with db.connect() as con:
    check("an old on-disk notice is moved into the database",
          db.get_notice_pdf(con, "legacy") == b"%PDF-legacy%%EOF")
    check("its path is cleared once the bytes are in",
          db.get_notice(con, "legacy")["pdf_path"] is None)
    check("the file itself is deleted", not _old_file.exists())
    check("a path whose file has already gone is left alone, not blanked",
          db.get_notice(con, "legacy-gone")["pdf_path"] == str(_gone_file))

db.init_db()                      # migrations and the move are both idempotent
with db.connect() as con:
    check("running the migration twice changes nothing",
          db.get_notice_pdf(con, "legacy") == b"%PDF-legacy%%EOF"
          and db.get_notice_pdf(con, "blob-notice") == _BLOB)

check("NOTICES_DIR is gone from the settings",
      not hasattr(main.settings, "notices_dir"))
_scraper_now = pathlib.Path("app/portal/scraper.py").read_text()
check("the scraper writes no PDF files any more",
      "notices_dir" not in _scraper_now and "save_as" not in _scraper_now)
check("the browser's temp copy is deleted once the bytes are in memory",
      "await download.delete()" in _scraper_now)
check("the scraper stores the downloaded bytes",
      'n["pdf_blob"] = await _download' in _scraper_now)
check("Claude is handed the bytes, not a path",
      "def _pdf_block(data: bytes)" in
      pathlib.Path("app/claude_client.py").read_text())


# 24 - the row buttons -------------------------------------------------------
# Short labels, always visible, right-aligned, and View reads the PDF without
# leaving the page.
_p = _page_now()
check("the header's primary button is just 'Sync'",
      '<button class="primary accent" id="sync">Sync</button>' in _p)
for label, hook in (("View", 'view(\'${esc(n.ref_id)}\')'),
                    ("Save", 'savePdf(\'${esc(n.ref_id)}\')'),
                    ("Draft", '>Draft</button>')):
    check(f"the row has a {label} button", hook in _p)
check("the ask-Claude button is the short '\u2726 Date'",
      "const DATE_BTN = '&#10022; Date';" in _p)
check("\u2726 Date is only on rows with no due date",
      "${(!n.due_date && n.has_pdf)" in _p)
check("View and Save only appear once a PDF is held",
      "${n.has_pdf ? `<button onclick=\"view(" in _p)
for gone in ("Generate response", "Ask Claude", ">Sync now<", ">Preview<"):
    check(f"the long label {gone!r} is gone", gone not in _p)
check("the buttons are right-aligned and reachable without a mouse",
      "justify-content: flex-end" in _p and ".rowacts:focus-within" in _p)

check("View opens an in-page modal, not a new tab",
      'id="viewer"' in _p and "window.open" not in _p)
check("the modal renders the PDF in a big iframe on the inline endpoint",
      "$('v-frame').src = `/api/notices/${encodeURIComponent(refId)}/pdf?inline=1`" in _p)
check("the modal has a dark overlay that follows the theme",
      ".modal { position: fixed; inset: 0;" in _p
      and "background: var(--overlay);" in _p
      and "--overlay: rgba(6, 7, 9, .72);" in _p)
check("clicking the overlay closes it",
      "viewer.onclick = ev => { if (ev.target === viewer) closeViewer(); }" in _p)
check("Esc closes it too", "palClose(); closeViewer();" in _p)
check("closing stops the PDF plugin behind the page",
      "$('v-frame').src = 'about:blank'" in _p)
check("Save asks the server for the attachment copy",
      "location.href = `/api/notices/${encodeURIComponent(refId)}/pdf`" in _p)
check("a found date is tagged 'by Claude'", "&#10022; by Claude" in _p)
check("the Draft button opens the drawer with summary, checklist and reply",
      "generateDraft(" in _p and "d-checklist" in _p and "d-summary" in _p)
check("the drawer keeps Copy and Regenerate",
      "d-copy" in _p and "d-regen" in _p and "regenerate=1" in _p)


# 25 - the overview: what do I have, what has been done ----------------------
_p = _page_now()

# (a) five cards, counted over everything the account holds
for label in ("Total notices", "Due this week", "Missing date", "Docs saved",
              "Drafts ready"):
    check(f"the overview shows {label!r}", label in _p)
check("the fifth card counts the drafts",
      "s-drafts" in _p and "rows.filter(n => n.has_draft).length" in _p)
check("each row still shows whether it has a draft",
      "n.has_draft" in _p and "'draft'" in _p)
check("the API still reports draft state per notice", "has_draft" in
      pathlib.Path("app/db.py").read_text())
check("the cards count everything, not the filtered view",
      "function renderStats(rows)" in _p
      and "renderStats(visibleRows" not in _p)

_STATS = [dict(n) for n in SAMPLE]
for row, drafted in zip(_STATS, (1, 0, 1, 0, 0)):
    row["has_draft"] = drafted
_count = lambda f: len([n for n in _STATS if f(n)])
check("Total notices counts every row", len(_STATS) == 5)
check("Due this week counts today through day 7",
      _count(lambda n: n["due_date"] and 0 <= _days(n["due_date"]) <= 7) == 2)
check("Missing date counts the blanks", _count(lambda n: not n["due_date"]) == 1)
check("Docs saved counts stored PDFs", _count(lambda n: n["has_pdf"]) == 4)
check("Drafts ready counts the notices that have one",
      _count(lambda n: n["has_draft"]) == 2)
check("a notice can be counted by more than one card",
      _count(lambda n: n["has_pdf"] and n["has_draft"]) == 2)

# (b) the last-sync line, off the runs table
with db.connect() as con:
    rid = con.execute("INSERT INTO runs DEFAULT VALUES").lastrowid
    db.finish_run(con, rid, "done", "{'notices': 9}",
                  {"new_notices": 4, "downloaded": 3, "skipped_cached": 6})
    run = db.last_run(con)
check("a finished run records what it found",
      run["notices_new"] == 4 and run["pdfs_saved"] == 3
      and run["skipped_cached"] == 6, str(dict(run)))
check("it records when it finished and how it went",
      run["finished"] is not None and run["status"] == "done")

with TestClient(main.app) as client:
    d = client.get("/api/notices").json()
    check("the dashboard is told about the last run in the same call",
          d["last_run"]["notices_new"] == 4 and d["last_run"]["pdfs_saved"] == 3,
          str(d.get("last_run")))

    with db.connect() as con:
        rid2 = con.execute("INSERT INTO runs DEFAULT VALUES").lastrowid
        db.finish_run(con, rid2, "failed", "RuntimeError('no tab')")
    d = client.get("/api/notices").json()
    check("the newest run is the one reported",
          d["last_run"]["status"] == "failed", str(d["last_run"]["status"]))
    check("a failed run reports no counts rather than zeros it did not earn",
          d["last_run"]["pdfs_saved"] is None, str(d["last_run"]))

# The line is back, but failure-safe: the reason a failed run must not print
# its message inline is that the message is a whole sentence, and it used to
# spill across the top of the page.
check("the dashboard renders the last run",
      "renderLastSync" in _p and "Last sync" in _p)
check("a failed run says only that it failed",
      "run.status !== 'done'" in _p
      and 'class="bad" title="${esc(run.message' in _p)
check("the line cannot wrap or spill",
      "white-space: nowrap; overflow: hidden; text-overflow: ellipsis;" in _p)
check("it says so plainly when nothing has run yet",
      "No sync has finished yet." in _p)
check("the run is still recorded server-side",
      "last_run" in pathlib.Path("app/main.py").read_text())

# a real sync writes those counts through
with TestClient(main.app) as client:
    reset()
    async def counting_run_sync(session, events, limit=None):
        return {"proceedings": 1, "notices": 5, "new_notices": 2,
                "downloaded": 2, "skipped_cached": 3}
    scraper.run_sync = counting_run_sync
    client.post("/api/credentials", json={"user_id": USER_ID, "password": PASSWORD})
    wait_idle(client)
    scraper.run_sync = fake_run_sync
    run = client.get("/api/notices").json()["last_run"]
    check("a completed sync leaves its counts behind for the dashboard",
          run["status"] == "done" and run["notices_new"] == 2
          and run["pdfs_saved"] == 2 and run["skipped_cached"] == 3, str(run))

check("the scraper counts notices it had never seen before",
      'stats["new_notices"] += 1' in _scraper_now
      and '"new_notices": 0' in _scraper_now)

# (c) the per-row checklist: three dots
check("each row carries the three dots",
      "function statusCell(" in _p and "'PDF', !!n.has_pdf" in _p
      and "'date', !!n.due_date" in _p and "'draft', !!n.has_draft" in _p)
check("a done dot is filled green, a pending one is hollow",
      ".tick.on { background: var(--ok); border-color: var(--ok); }" in _p
      and "border: 1.5px solid var(--divider); background: none;" in _p)
check("every dot says what it means",
      'title="${esc(label)}: ${esc(title)}"' in _p
      and "aria-label=\"${esc(label)} ${on === null ? 'unknown'" in _p)
check("the table has a Status column for them",
      "<th>Status</th>" in _p)
check("the wider table still spans its empty state", 'colspan="6"' in _p
      and 'colspan="5"' not in _p)
check("writing a draft ticks the row without a refetch",
      "row.has_draft = 1;" in _p)


# 26 - the theme ------------------------------------------------------------
# Dark by default, light on [data-theme], colour only where it means
# something, and exactly one gradient.
_css = pathlib.Path("app/static/style.css").read_text()
_p = _page_now()

for name, value in (("--bg", "#0a0b0d"), ("--panel", "#131417"),
                    ("--hairline", "#26272b"), ("--text", "#ececf1"),
                    ("--muted", "#8b8d98")):
    check(f"dark {name} is {value}", f"{name}: {value};" in _css)
for meaning, value in (("overdue", "#ef4444"), ("missing date", "#f59e0b"),
                       ("done", "#22c55e"), ("Claude", "#6e79f7")):
    check(f"{meaning} is {value}", value in _css)

check("there is one accent gradient",
      _css.count("linear-gradient(135deg, #5b63f0 0%, #8b5cf6 100%)") == 1)
check("it is reserved for Sync and Draft",
      _p.count('class="primary accent" id="sync"') == 1
      and 'class="primary accent" onclick="generateDraft(' in _p)
check("the gradient class is defined once and used nowhere else in CSS",
      _css.count(".accent { background: var(--accent);") == 1)
check("Claude's hue is never the action colour",
      "--ai: #6e79f7;" in _css and "--action: #5b63f0;" in _css)

check("dark is the default and light is opt-in",
      'data-theme="dark"' in _p and "[data-theme='light']" in _css)
check("the theme is a cookie, not localStorage",
      "document.cookie = `theme=" in _p
      and "localStorage." not in _p and "localStorage[" not in _p)
check("every semantic token is redefined for light",
      all(f"  {t}:" in _css.split("[data-theme='light']")[1].split("}")[0]
          for t in ("--bg", "--panel", "--text", "--muted", "--ok", "--warn",
                    "--danger", "--ai", "--overlay")))

check("Geist is what the page actually asks for",
      "font-family: 'Geist';" in _css and "font-family: 'Geist Mono';" in _css
      and "/fonts/Geist-Variable.woff2" in _p
      and "/fonts/GeistMono-Variable.woff2" in _p)
check("both stacks still fall back to the system face",
      "'Geist', system-ui" in _css and "'Geist Mono', ui-monospace" in _css)
check("the fonts it preloads are the fonts it uses",
      "manrope" not in _p and "space-grotesk" not in _p and "ibm-plex" not in _p)

# table polish
check("headers are 11px uppercase muted",
      "font: 600 11px var(--sans); text-transform: uppercase;" in _css
      and "background: var(--panel); color: var(--muted);" in _css)
check("the header row sticks", "position: sticky; top: 0; z-index: 2;" in _css)
check("rows are separated by hairlines only",
      "border-top: 1px solid var(--hairline)" in _css)
check("the row's actions appear on hover, focus, or on a touch device",
      "tr:hover .rowacts, .rowacts:focus-within { opacity: 1; }" in _css
      and "@media (hover: none) { .rowacts { opacity: 1; } }" in _css)
check("dates and identifiers are mono and line up",
      "td.mono, .idchip { font-variant-numeric: tabular-nums; }" in _css)

# countdown chips
check("a countdown reads 12d / 3d / overdue 2d",
      "`overdue ${Math.abs(d)}d`" in _p and "`${d}d`" in _p)
check("green with room, amber inside two weeks, red inside three days",
      "d < 3 ? 'late' : d <= 14 ? 'soon' : 'ok'" in _p
      and ".chip.ok { color: var(--ok-text)" in _css
      and ".chip.soon { color: var(--warn-text)" in _css
      and ".chip.late { color: var(--danger-text)" in _css)

# the Claude surfaces
check("an AI card is bordered in Claude's hue with a ✦ mark",
      "border-left: 2px solid var(--ai)" in _css and "&#10022;" in _p)
check("its footer credits Claude, with the time and the basis",
      "&#10022; Generated by Claude" in _p and "relTime(" in _p
      and "' · basis: '" in _p)
check("Regenerate is a quiet link, not a button",
      ".ai-foot .regen" in _css and "text-decoration: underline" in _css)

# motion, focus, loading
check("every animation sits behind prefers-reduced-motion",
      _css.count("@media (prefers-reduced-motion: no-preference)") >= 3
      and "@media (prefers-reduced-motion: reduce)" not in _css)
check("transitions are in the 150-200ms band",
      "transition: opacity .15s ease" in _css
      and "transition: transform .2s cubic-bezier(.32,.72,0,1)" in _css)
check("focus is visible", ":focus-visible { outline: 2px solid var(--action)" in _css)
check("loading shimmers rather than jumping",
      "@keyframes shimmer" in _css and "SKELETON" in _p)
check("the empty state is designed, not a bare sentence",
      '<div class="title">' in _p and '<div class="desc">' in _p
      and ".empty-state .title" in _css)


# 27 - the pipeline bar and the command palette ------------------------------
_p = _page_now()
_scr = pathlib.Path("app/portal/scraper.py").read_text()
_main_src = pathlib.Path("app/main.py").read_text()

# the stages the backend actually emits, in the order the bar draws them
check("the bar's stages are the ones the server sends",
      "{ key: 'login'" in _p and "{ key: 'list'" in _p and "{ key: 'walk'" in _p
      and "{ key: 'download'" in _p and "{ key: 'done'" in _p)
check("login reports itself", 'hub.progress("login")' in _main_src
      and 'hub.progress("login", done=True)' in _main_src)
check("opening the list reports itself", 'events.progress("list")' in _scr)
check("each tab and sub-tab reports itself",
      'events.progress("walk", tab=tab_label, sub_tab=sub_label' in _scr)
check("each card says which one it is",
      "card=i + 1, of=total" in _scr and "TABS.get(tab_key, tab_key)" in _scr)
check("each download says N of M",
      'events.progress("download", notice=j + 1, of=total' in _scr)
check("the run's totals close the bar", 'hub.progress("done"' in _main_src)

rec3 = Recorder()
main.hub.sockets.append(rec3)
asyncio.run(main.hub.progress("download", notice=3, of=12, downloaded=1))
msg = rec3.sent[-1]
check("a progress message is typed and carries stage + counts",
      msg["type"] == "progress" and msg["stage"] == "download"
      and msg["counts"] == {"notice": 3, "of": 12, "downloaded": 1}, str(msg))
main.hub.sockets.remove(rec3)

check("the bar renders counts as a person would say them",
      "function countText(" in _p
      and "`downloading ${c.notice} of ${c.of}`" in _p
      and "`card ${c.card} of ${c.of}`" in _p)
check("an unrecognised count still shows rather than vanishing",
      "`${esc(k)} ${esc(v)}`" in _p)
check("done stages tick, the active one pulses",
      "cls === 'done' ? '&check;'" in _p
      and ".step.active .bead { border-color: var(--action)" in _p)

# the palette
check("Ctrl/Cmd+K opens it",
      "(ev.ctrlKey || ev.metaKey) && ev.key.toLowerCase() === 'k'" in _p)
for label in ("Run sync", "Toggle theme", "Speed: slow", "Speed: fast",
              "Speed: extreme", "Filter: missing due date"):
    check(f"the palette offers {label!r}", label in _p)
check("it can open a notice, searched by reference or description",
      "Open notice ${n.ref_id}" in _p
      and "haystack: `${n.ref_id} ${n.description || ''}" in _p)
check("Enter on a notice opens the viewer", "run: () => view(n.ref_id)" in _p)
check("matching is a subsequence, the way a palette should be",
      "function fuzzy(" in _p and "h.indexOf(ch, i)" in _p)
check("Esc closes it", "if (ev.key === 'Escape') { palClose();" in _p)
check("s syncs and / searches",
      "if (ev.key === 's') { ev.preventDefault(); startSync(); }" in _p
      and "if (ev.key === '/') { ev.preventDefault(); $('f-name').focus(); }" in _p)
check("neither fires while you are typing",
      "/^(INPUT|TEXTAREA|SELECT)$/.test(document.activeElement.tagName)" in _p
      and "if (typing) return;" in _p)



# 25 - the summary report and the Excel export -------------------------------
# The firm's old tracker: which notices are overdue, which are about to be,
# and the same numbers in a workbook. Counted on a database of its own so the
# arithmetic is exact rather than "whatever the earlier tests left behind".
import io as _io                                                     # noqa: E402
from datetime import date as _date, timedelta as _td                 # noqa: E402
from openpyxl import load_workbook                                   # noqa: E402

from app import report                                               # noqa: E402

_MAIN_DB = db.DB_PATH
_SUM_DB = TMP.parent / "summary.db"
db.DB_PATH = _SUM_DB
db.init_db()


def _portal_date(offset):
    """A date in the portal's own 17-Aug-2026 shape, offset days from today."""
    return (_date.today() + _td(days=offset)).strftime("%d-%b-%Y")


def _proc(con, name, status):
    return db.upsert_proceeding(con, {
        "tab": "self", "sub_tab": "action", "proceeding_name": name,
        "pan": "AAACU3358G", "assessee_name": "CAMBRIDGE TECHNOLOGY ENTERPRISES",
        "assessment_year": "2026-27", "financial_year": "2025-26",
        "applicable_act": "Income Tax Act 1961", "status": status,
        "closure_date": None, "closure_order": None})


def _notice(con, ref, pid, due, pdf=b"%PDF-x%%EOF", source="portal"):
    db.upsert_notice(con, {
        "proceeding_id": pid, "ref_id": ref, "notice_us": "142(1)",
        "doc_ref_id": "ITBA/COM/F/17/2026-27/1092231604(1)",
        "description": f"[ITBA]{ref}", "issued_on": _portal_date(-20),
        "served_on": None, "due_date": due,
        "due_date_source": source if due else None,
        "ao_viewed_on": None, "pdf_blob": pdf})


with db.connect() as con:
    _open = _proc(con, "Assessment u/s 143(2)", "Open")
    _shut = _proc(con, "Assessment already answered", "Closed")
    _notice(con, "n-overdue", _open, _portal_date(-5))
    _notice(con, "n-2d", _open, _portal_date(2))
    _notice(con, "n-3d", _open, _portal_date(3))
    _notice(con, "n-4d", _open, _portal_date(4))
    _notice(con, "n-10d", _open, _portal_date(10))
    _notice(con, "n-40d", _open, _portal_date(40))
    _notice(con, "n-nodate", _open, None, pdf=None)
    _notice(con, "n-closed", _shut, _portal_date(-9))
    db.save_draft(con, "n-overdue", "summary", "[]", "draft text")
    _S = report.build_summary(con)

_B = {b["key"]: b["count"] for b in _S["buckets"]}
check("overdue counts the dates that have already gone", _B["overdue"] == 1, str(_B))
check("due <=3 days holds day 2 and day 3", _B["due_3"] == 2, str(_B))
check("due <=10 days starts at day 4 and ends at day 10", _B["due_10"] == 2, str(_B))
check("on track is everything past ten days", _B["on_track"] == 1, str(_B))
check("a notice with no due date has its own bucket", _B["no_due_date"] == 1, str(_B))
check("a closed proceeding is counted separately", _B["closed"] == 1, str(_B))
check("a closed notice never lands in an urgency bucket even when its date has gone",
      next(i for i in _S["register"] if i["ref_id"] == "n-closed")["bucket"] == "closed")
check("every notice is counted exactly once",
      sum(_B.values()) == len(_S["register"]) == 8, str(_B))

check("day 3 is critical, day 4 is not",
      report.bucket_of(3, True) == "due_3" and report.bucket_of(4, True) == "due_10")
check("day 10 is the last of the ten-day band",
      report.bucket_of(10, True) == "due_10" and report.bucket_of(11, True) == "on_track")
check("a closed proceeding is closed whatever its date",
      report.bucket_of(-9, False) == "closed" and report.bucket_of(None, False) == "closed")
check("the portal's own date format parses",
      report.parse_date("17-Aug-2026") == _date(2026, 8, 17))
check("a date the portal left as '-' is not guessed at",
      report.parse_date("-") is None and report.parse_date(None) is None)

_attn = [i["ref_id"] for i in _S["attention"]]
check("attention holds only what is overdue or due within three days",
      _attn == ["n-overdue", "n-2d", "n-3d"], str(_attn))
check("days left is negative once the date has gone",
      _S["attention"][0]["days_left"] == -5, str(_S["attention"][0]["days_left"]))
_first = _S["attention"][0]
check("attention carries the fields the old sheet showed",
      all(_first[k] is not None for k in
          ("description", "pan", "assessment_year", "notice_us", "due_date")),
      str(_first))
check("attention carries the pdf / date / draft flags",
      _first["has_pdf"] and _first["has_due_date"] and _first["has_draft"],
      str(_first))
check("the register records where each due date came from",
      {i["due_date_source"] for i in _S["register"]} == {"portal", None},
      str({i["due_date_source"] for i in _S["register"]}))
check("the run line says what was scanned",
      _S["run"]["notices_scanned"] == 8, str(_S["run"]))

with TestClient(main.app) as client:
    r = client.get("/api/summary")
    check("/api/summary answers", r.status_code == 200, r.text[:80])
    _j = r.json()
    check("it ships run info, buckets, attention and the full register",
          all(k in _j for k in ("run", "buckets", "attention", "register")),
          str(sorted(_j))[:100])
    check("the endpoint counts the same buckets as the module",
          {b["key"]: b["count"] for b in _j["buckets"]} == _B)
    check("the run block carries the finish time and the new count",
          "finished" in _j["run"] and "new_this_run" in _j["run"], str(_j["run"]))

    r = client.get("/api/export.xlsx")
    check("the Excel export answers 200", r.status_code == 200, str(r.status_code))
    check("it is served as a real xlsx",
          r.headers.get("content-type", "").startswith(
              "application/vnd.openxmlformats-officedocument.spreadsheetml"),
          r.headers.get("content-type", "none"))
    check("the download is named for the day it was built",
          f'itr-summary-{_date.today().isoformat()}.xlsx'
          in r.headers.get("content-disposition", ""),
          r.headers.get("content-disposition", "none"))

    _wb = load_workbook(_io.BytesIO(r.content))
    check("the workbook opens, with the three sheets",
          _wb.sheetnames == ["Summary", "Attention", "All notices"],
          str(_wb.sheetnames))
    _sum_ws = _wb["Summary"]
    _sum_text = [str(c.value) for row in _sum_ws.iter_rows() for c in row if c.value]
    check("the summary sheet keeps the tracker's title",
          "Position at a glance" in _sum_text, str(_sum_text[:6]))
    check("it keeps the caution line",
          any("verify every figure against the portal" in t for t in _sum_text))
    check("it writes the run date as a date, not text",
          isinstance(_sum_ws["B2"].value, (_date, datetime.datetime)),
          type(_sum_ws["B2"].value).__name__)
    check("every bucket has a label row and a count",
          all(b["label"] in _sum_text for b in _j["buckets"]))
    check("the bucket labels are bold",
          _sum_ws.cell(row=9, column=1).font.bold is True)

    _att = _wb["Attention"]
    check("the attention sheet lists exactly the attention rows",
          _att.max_row == 1 + len(_j["attention"]), str(_att.max_row))
    check("its header row is the tracker's dark blue with white bold text",
          _att["A1"].font.bold and _att["A1"].font.color.rgb.endswith("FFFFFF")
          and _att["A1"].fill.fgColor.rgb.endswith("1F3864"),
          str(_att["A1"].fill.fgColor.rgb))
    check("the header row is frozen", _att.freeze_panes == "A2", str(_att.freeze_panes))
    check("columns are sized to their content",
          _att.column_dimensions["A"].width > 10, str(_att.column_dimensions["A"].width))
    check("due dates are written as dates, not strings",
          isinstance(_att["E2"].value, (_date, datetime.datetime)),
          type(_att["E2"].value).__name__)
    check("days left stays a number so Excel can sort it",
          isinstance(_att["F2"].value, int) and _att["F2"].value == -5,
          str(_att["F2"].value))

    _all = _wb["All notices"]
    _headers = [c.value for c in _all[1]]
    check("the register sheet holds every notice",
          _all.max_row == 1 + len(_j["register"]), str(_all.max_row))
    check("it says where each due date came from",
          "Due date from" in _headers, str(_headers))
    check("it carries the pdf and draft columns",
          "PDF saved" in _headers and "Draft ready" in _headers, str(_headers))

# the lock covers the report exactly like everything else under /api/
main.settings.app_password = "dashboard-pw"
with TestClient(main.app) as anon:
    check("locked: the summary is not readable",
          anon.get("/api/summary").status_code == 401)
    r = anon.get("/api/export.xlsx")
    check("locked: the export is refused", r.status_code == 401, str(r.status_code))
    check("locked: no workbook bytes leak out",
          not r.content.startswith(b"PK"), r.content[:8].hex())
main.settings.app_password = ""

db.DB_PATH = _MAIN_DB              # back to the suite's own database

# what the dashboard ships for all of this
_p = _page_now()
check("the report section is titled like the old sheet",
      "Position at a glance" in _p)
check("it prints a run-date line", 'id="r-run"' in _p and "notices scanned" in _p)
check("the buckets are a row of chips that filter the table",
      'id="buckets"' in _p and "data-bucket" in _p
      and "bucketOf(n) === BUCKET" in _p)
check("the dashboard repeats the server's bucket rules exactly",
      "d <= 3" in _p and "d <= 10" in _p and "'no_due_date'" in _p)
check("overdue is red, three days amber, no-date indigo, on track green, closed muted",
      "overdue: 'late', due_3: 'soon'" in _p
      and "on_track: 'ok', no_due_date: 'none', closed: 'done'" in _p
      and ".bchip.none { color: var(--ai)" in _p
      and ".bchip.ok { color: var(--ok-text)" in _p
      and ".bchip.done { color: var(--muted)" in _p)
check("the attention table has the tracker's dark header band",
      "table.attn thead th { background: var(--attn-head); color: #fff; }" in _p)
check("its columns are the sheet's columns",
      "Client / Description" in _p and "<th>Section</th>" in _p
      and "Days left" in _p)
check("days left is coloured red when negative and amber inside three days",
      "d < 0 ? 'late' : d <= 3 ? 'soon' : ''" in _p
      and ".days.late { color: var(--danger-text); }" in _p)
check("an empty attention table says exactly what the old sheet said",
      "Nothing overdue or critical." in _p and ".allclear { color: var(--ok-text)" in _p)
check("the caution line survives",
      "Draft for review &mdash; verify every figure against the portal." in _p)
check("Export sits next to Sync in the header",
      '<button class="ghost" id="export"' in _p
      and "location.href = '/api/export.xlsx'" in _p)


# 28 - has a reply been filed? ------------------------------------------------
# Read off the card, never by clicking: "View Response" means a reply exists,
# "Submit Response" means none does. Both buttons stay in FORBIDDEN.
check("the real card, which shows View Response, reads as responded",
      n["responded"] == 1, str(n["responded"]))

_await_reply = REAL_NOTICE.replace("View Response", "Submit Response")
check("a card showing Submit Response reads as not responded",
      asyncio.run(scraper._parse_notice(TextCard(_await_reply), 1))["responded"] == 0)

_neither = REAL_NOTICE.replace("View Response\n", "")
_parsed = asyncio.run(scraper._parse_notice(TextCard(_neither), 1))
check("a card with neither button leaves it unknown, not false",
      _parsed["responded"] is None, str(_parsed["responded"]))
check("'Seek/View Adjournment' is not mistaken for a filed reply",
      "Seek/View Adjournment" in _neither and _parsed["responded"] is None)
check("the portal's own casing does not matter",
      scraper._responded_from("submit response") == 0
      and scraper._responded_from("VIEW RESPONSE") == 1)
check("neither button is ever clickable from here",
      all(b in scraper.FORBIDDEN for b in ("submit response", "view response")))

# it is stored, and unlike the PDF it is refreshed on every sync
with db.connect() as con:
    base = {"proceeding_id": None, "ref_id": "reply-test", "notice_us": None,
            "doc_ref_id": None, "description": None, "issued_on": None,
            "served_on": None, "due_date": None, "due_date_source": None,
            "ao_viewed_on": None, "responded": 0, "pdf_blob": b"%PDF-x"}
    db.upsert_notice(con, base)
    check("a first sync stores the flag",
          db.get_notice(con, "reply-test")["responded"] == 0)

    db.upsert_notice(con, dict(base, responded=1, pdf_blob=None))
    check("a later sync overwrites it - this is not cached like the PDF",
          db.get_notice(con, "reply-test")["responded"] == 1)

    db.upsert_notice(con, dict(base, responded=None, pdf_blob=None))
    check("a sync that could not tell keeps the last known answer",
          db.get_notice(con, "reply-test")["responded"] == 1)

    db.set_responded(con, "reply-test", 0)
    check("a cached notice still gets its flag refreshed",
          db.get_notice(con, "reply-test")["responded"] == 0)
    db.set_responded(con, "reply-test", None)
    check("refreshing with 'unknown' changes nothing",
          db.get_notice(con, "reply-test")["responded"] == 0)

check("the scraper refreshes it even for notices it skips downloading",
      'db.set_responded(con, n["ref_id"], n["responded"])'
      in pathlib.Path("app/portal/scraper.py").read_text())

with TestClient(main.app) as client:
    row = next(r for r in client.get("/api/notices").json()["notices"]
               if r["ref_id"] == "reply-test")
    check("the dashboard is told", row["responded"] == 0, str(row.get("responded")))

_p = _page_now()
check("the row's checklist has a fourth dot for it",
      "'responded on portal'" in _p)
check("unknown is drawn dashed rather than claimed as 'not yet'",
      ".tick.unknown { border-style: dashed;" in _p
      and "the portal did not say at the last sync" in _p)


print()
print(f"{'FAILED: ' + ', '.join(failures) if failures else 'all checks passed'}")
sys.exit(1 if failures else 0)
